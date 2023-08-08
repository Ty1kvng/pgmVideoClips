# Import necessary modules
from importlib.resources import path
import inspect
from modulefinder import packagePathMap
from moviepy.editor import VideoFileClip
import shutil
import openpyxl
import os

# Get input data of clip name and timestamp locations
#####
# Get the absolute path of the spreadsheet
loc = input("Enter absolute path of spreadsheet: ")
# Open the workbook
wb_obj = openpyxl.load_workbook(loc, data_only=True)
# Get the active sheet in the workbook
sheet_obj = wb_obj.active
# Starting row is 3 after header, columns stay the same
rowCount = 3

## Retrieving Clip ##
# Load the video
vidName = input("Enter video name and extension:  ")


# Loop through rows in the spreadsheet
for i in range(rowCount, sheet_obj.max_row+1):
    rowCount = i

    # Set start and end times of timestamp
    tStart = sheet_obj.cell(row=rowCount, column=4)
    tEnd = sheet_obj.cell(row=rowCount, column=5)

    # Set location of clip flags
    cfloc = sheet_obj.cell(row=rowCount, column=1)

    # If clip is not flagged as completed and has valid timestamps
    if ((sheet_obj.cell(row=rowCount, column=1).value == "N") and (sheet_obj.cell(row=rowCount, column=6).value != "00:00:00.000")):

        # Load the video clip
        clip = VideoFileClip(
            os.path.abspath('Main/') + "\\" + vidName)

        # Select only the video between the specified timestamps as strings
        clip = clip.subclip(str(tStart.value), str(tEnd.value))
        
        # Create the filename for the new clip
        clipName = (str(sheet_obj.cell(row=rowCount, column=2).value) +
                    "_" + (str(tStart.value))[:-4].replace(":", ".") + ".mp4")
        
        # Write the new clip to a file
        clip.write_videofile(clipName)
        
        # Update the filename in the spreadsheet
        fileNameLoc = sheet_obj.cell(row=rowCount, column=2)
        fileNameLoc.value = clipName

        #####################
        # Move the clip to the Clips folder
        pgmsource = (os.path.dirname(os.path.abspath(
            inspect.getfile(inspect.currentframe())))) + "\\"

        destination = os.path.abspath('Clips/')
        dest = shutil.move(pgmsource + clipName, destination,
                           copy_function=shutil.copytree)

        # Flag clip as completed
        cfloc.value = "Y"
        # Save the updated spreadsheet
        wb_obj.save(loc)
    else:
        # Continue to the next row in the spreadsheet
        continue

# Print message when finished processing spreadsheet
print("End of spreadsheet!")
