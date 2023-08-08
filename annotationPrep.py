import pandas as pd
import os
from openpyxl import load_workbook

file_path = input("Enter path to Excel file: ")
file_name = os.path.basename(file_path)
sheet_name = None

# Load the workbook and get the sheet names
wb = load_workbook(file_path, read_only=True)
sheet_names = wb.sheetnames

# If there is only one sheet, set sheet_name to that sheet
if len(sheet_names) == 1:
    sheet_name = sheet_names[0]
else:
    # If there are multiple sheets, prompt the user to choose one
    print("Sheet names:")
    for i, name in enumerate(sheet_names):
        print(f"{i+1}. {name}")
    choice = input("Enter the number of the sheet you want to use: ")
    sheet_name = sheet_names[int(choice)-1]

# Load the sheet into a dataframe using the chosen sheet_name
df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')

# Insert two columns at the beginning
df.insert(0, 'New Column 1', '')
df.insert(1, 'New Column 2', '')

# Fill in the required data
df.iloc[1, 0] = 'Clip Created (Y/N)'
df.iloc[1, 1] = 'Clip Filename'
df.iloc[2, 0] = 'N'
df.iloc[2, 1] = f'={"_".join(file_name.split(".")[0][:6])}&IF(K3<10,"0"&K3,K3)&J3'

# Write the updated dataframe to the original Excel file
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = writer.book  # Load existing file
writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
writer.book.active = writer.book.sheetnames.index(file_name)  # Set the active sheet
writer.save()

# Open the updated Excel file
os.startfile(file_path)
