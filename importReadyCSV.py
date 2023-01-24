# Import everything needed to edit video clips
from __future__ import annotations
import openpyxl
from openpyxl import Workbook
import csv
import os

# Getting input data of clip name and timestamp locations from original workbook
#####
# Location of spreadsheet
loc = input("Enter absolute path of spreadsheet1: ")
# Open orignal workbook
wb_obj = openpyxl.load_workbook(loc, data_only=True)
sheet_obj = wb_obj.active

# Open new workbook object
wb_obj2 = Workbook()
sheet_obj2 = wb_obj2.active

# Starting row is 3 after header (original wb), columns stay the same
rowCount = 3

# Add titles to first row of new sheet
sheet_obj2.cell(row=1, column=1).value = "Clip Filename"
sheet_obj2.cell(row=1, column=2).value = "Start"
sheet_obj2.cell(row=1, column=3).value = "Duration"
sheet_obj2.cell(
    row=1, column=4).value = "Combined_Tag (includes Label and Tags)"

# Iterate through the orignal workbook
for i in range(rowCount, sheet_obj.max_row + 1):
    rowCount = i

    # Check if output worksheet
    if ((sheet_obj.cell(row=rowCount, column=1).value == "Y") and (sheet_obj.cell(row=rowCount, column=6).value != "00:00:00.000")):

        # Transfer clipfile name from orignal to new workbook
        sheet_obj2.cell(
            row=rowCount-1, column=1).value = str(sheet_obj.cell(row=rowCount, column=2).value)
        # Transfer start times from original to new workbook
        sheet_obj2.cell(
            row=rowCount-1, column=2).value = str(sheet_obj.cell(row=rowCount, column=4).value)
        # Transfer duration from original to new workbook
        sheet_obj2.cell(
            row=rowCount-1, column=3).value = str(sheet_obj.cell(row=rowCount, column=6).value)

        # Create a search for all each X value in each column
        # Combine title of row location (tag) where each x appears. Place in 4th row
        # sheet_obj2.cell(row=rowCount-1, column=4).value = (sheet_obj.cell(row=rowCount, column=7).value)
        tempStr = ""
        for j in range(1, sheet_obj.max_column + 1):
                # Skip if tag is same as label
                if(sheet_obj.cell(row=rowCount, column=j).value == 'X' and (sheet_obj.cell(row=2, column=j).value == (sheet_obj.cell(row=rowCount, column=7).value))):
                    continue
                if(sheet_obj.cell(row=rowCount, column=j).value == 'X'):
                    tempStr = tempStr + ";" + str(sheet_obj.cell(row=2, column=j).value)

        # Add the temp string of tags to the label
        sheet_obj2.cell(row=rowCount-1, column=4).value = (
            sheet_obj.cell(row=rowCount, column=7).value) + tempStr
        
    else:
        continue
    
# Save new workbook
wb_obj2.save(filename="temp.xlsx")

#### Convert to CSV ####

# open given workbook 
# and store in excel object 
excel = openpyxl.load_workbook("temp.xlsx")
  
# select the active sheet
sheet = excel.active
  
# writer object is created
col = csv.writer(open("_import_ready_annotations.csv",
                      'w', 
                      newline=""))
  
# writing the data in csv file
for r in sheet.rows:
    # row by row write 
    # operation is perform
    col.writerow([str(cell.value) for cell in r])

os.remove("temp.xlsx")
print("New CSV Created!")
