# Importing necessary modules
from importlib.resources import path
import inspect
from modulefinder import packagePathMap
from moviepy.editor import VideoFileClip
import shutil
import openpyxl
import os

# Getting input data of clip name and timestamp locations
#####
# Location of spreadsheet
loc = input("Enter absolute path of spreadsheet: ")
# Open workbook
wb_obj = openpyxl.load_workbook(loc, data_only=True)
sheet_obj = wb_obj.active
# Starting row is 3 after header, columns stay the same
rowCount = 3

## Retrieving Clip ##

# getting the name of the video file from the user
vidName = input("Enter video name and extension:  ")

# looping through the spreadsheet
for i in range(rowCount, sheet_obj.max_row+1):
    rowCount = i

    # Set start and end times of timestamp
    tStart = sheet_obj.cell(row=rowCount, column=4)
    tEnd = sheet_obj.cell(row=rowCount, column=5)

    # Set location of clip flags
    cfloc = sheet_obj.cell(row=rowCount, column=1)

    # checking if clip flag is "N" and timestamp is not "00:00:00.000"
    if ((sheet_obj.cell(row=rowCount, column=1).value == "N") and (sheet_obj.cell(row=rowCount, column=6).value != "00:00:00.000")):

        # opening the video file using moviepy
        clip = VideoFileClip(
            os.path.abspath('Main/') + "\\" + vidName)

        # getting only video between timestamp 1 and 2 as strings
        clip = clip.subclip(str(tStart.value), str(tEnd.value))

        # Saving the clip with a new name
        clipName = (str(sheet_obj.cell(row=rowCount, column=2).value) +
                    "_" + (str(tStart.value))[:-4].replace(":", ".") + ".mp4")
        clip.write_videofile(clipName)

        # updating the clip name in the spreadsheet
        fileNameLoc = sheet_obj.cell(row=rowCount, column=2)
        fileNameLoc.value = clipName

        #####################
        # Move the clip to correct folder
        # getting the current directory of the program
        pgmsource = (os.path.dirname(os.path.abspath(
            inspect.getfile(inspect.currentframe())))) + "\\"

        # setting the destination folder
        destination = os.path.abspath('Clips/')

        # moving the clip to the destination folder
        dest = shutil.move(pgmsource + clipName, destination,
                           copy_function=shutil.copytree)

        # Change Clip Flag to "Y"
        cfloc.value = "Y"
        wb_obj.save(loc)
    else:
        continue

print("End of spreedsheet!")
