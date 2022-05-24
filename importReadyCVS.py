# Import everything needed to edit video clips
import openpyxl
import os

# Getting input data of clip name and timestamp locations from original workbook
#####
# Location of spreadsheet
loc = input("Enter absolute path of spreadsheet1: ")
# Open orignal workbook
wb_obj = openpyxl.load_workbook(loc, data_only=True)
sheet_obj = wb_obj.active

# Open new workbook object
wb_obj2 = openpyxl.load_workbook(loc, data_only=True)
sheet_obj2 = wb_obj2.active

# Starting row is 3 after header (original wb), columns stay the same
rowCount = 3

# Add titles to first row of new sheet
sheet_obj2.cell(row=1, column=1).value = "Clip Filename"
sheet_obj2.cell(row=1, column=2).value = "Start"
sheet_obj2.cell(row=1, column=3).value = "Duration"
sheet_obj2.cell(row=1, column=4).value = "Combined_Tag (includes Label and Tags)"

# Iterate through the orignal workbook
for i in range(rowCount, sheet_obj.max_+1):
    rowCount = i

    #Check if output worksheet
    if (sheet_obj.cell(row=rowCount, column=1).value == "N"):

        # Transfer clipfile name from orignal to new workbook
        sheet_obj2.cell(row=rowCount-1, column=1).value = sheet_obj.cell(row=rowCount, column=2).value
        # Transfer start times from original to new workbook
        sheet_obj2.cell(row=rowCount-1, column=2).value = sheet_obj.cell(row=rowCount, column=4).value
        # Transfer duration from original to new workbook
        sheet_obj2.cell(row=rowCount-1, column=3).value = sheet_obj.cell(row=rowCount, column=6).value
        
        
        ## Create a search for all each X value in column
        # Combine title of row location (tag) where each x appears. Place in 4th row
        sheet_obj2.cell(row=rowCount-1, column=4).value

        #Save new workbook
        wb_obj2.save(loc)
    else:
        continue

print("End of spreedsheet!")